import asyncio, openpyxl, time
import xl2web

from openpyxl import load_workbook

# Excel to Web, Example Class
class Xl2Web_c(xl2web.Xl2WebBase):
  # miscellanous helper
  def _pack_data(self, fullname):
    # reopen the worksheet all over again
    wb = load_workbook(self.xlsx, data_only=True)
    ws = wb["Example"]

    month_09 = {
      "budget": ws['E3'].value,
      "actual": ws['E4'].value,
      "gap"   : ws['E5'].value
     }

    month_10 = {
      "budget": ws['G3'].value,
      "actual": ws['G4'].value,
      "gap"   : ws['G5'].value
     }

    return {
      "timestamp" : time.ctime(),
      "file"      : fullname,
      "month_09"  : month_09,
      "month_10"  : month_10
     }

  def _dump_data(self, data):
    print("Timestamp     : %s" % data["timestamp"])
    print("File Modified : %s" % data["file"])
    print("September     : %s" % data["month_09"])
    print("October       : %s" % data["month_10"])
    print()

# Program Entry Point
example = Xl2Web_c(
  '/home/epsi/awatch/code-02-enh', 'test-c.xlsx',
  'localhost', 8765)
  
try:
  asyncio.run(example.main())
except KeyboardInterrupt:
  print('Goodbye!')

