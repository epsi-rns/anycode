import asyncio, openpyxl, time
import xl2web

from openpyxl import load_workbook

# Excel to Web, Example Class
class Xl2Web_d(xl2web.Xl2WebBase):
  # miscellanous helper
  def _pack_data(self, fullname):
    # reopen the worksheet all over again
    wb = load_workbook(self.xlsx, data_only=True)
    ws = wb["Example"]

    month_09 = {
      "target": ws['D3'].value,
      "actual": ws['E3'].value,
      "miss"  : ws['F3'].value,
      "remain": ws['G3'].value
     }

    month_10 = {
      "target": ws['D4'].value,
      "actual": ws['E4'].value,
      "miss"  : ws['F4'].value,
      "remain": ws['G4'].value
     }

    month_11 = {
      "target": ws['D5'].value,
      "actual": ws['E5'].value,
      "miss"  : ws['F5'].value,
      "remain": ws['G5'].value
     }

    total = {
      "target": ws['D6'].value,
      "actual": ws['E6'].value,
      "miss"  : ws['F6'].value,
      "remain": ws['G6'].value
     }

    return {
      "timestamp" : time.ctime(),
      "file"      : fullname,
      "month_09"  : month_09,
      "month_10"  : month_10,
      "month_11"  : month_11,
      "total"     : total
     }

  def _dump_data(self, data):
    print("Timestamp     : %s" % data["timestamp"])
    print("File Modified : %s" % data["file"])
    print("September     : %s" % data["month_09"])
    print("October       : %s" % data["month_10"])
    print("November      : %s" % data["month_11"])
    print("Total         : %s" % data["total"])
    print()

# Program Entry Point
example = Xl2Web_d(
  '/home/epsi/awatch/code-02-enh', 'test-d.xlsx',
  'localhost', 8767)
  
try:
  asyncio.run(example.main())
except KeyboardInterrupt:
  print('Goodbye!')

